"""
DEX JR XLSX CONVERTER — v1.0
Converts .xlsx files to .txt for corpus ingestion.

Usage:
  python dex-xlsx.py "C:\path\to\file.xlsx"
  python dex-xlsx.py "C:\path\to\folder" --all
  python dex-xlsx.py "C:\path\to\file.xlsx" --ingest
  python dex-xlsx.py "C:\path\to\folder" --all --ingest
  python dex-xlsx.py "C:\path\to\file.xlsx" --save "C:\output\folder"

Modes:
  (default)   Convert and display
  --all       Convert all .xlsx files in a folder
  --ingest    Save converted text directly to canon folder
  --save      Save converted text to specified folder
  --preview   Show first 50 rows per sheet only

Each sheet becomes a section in the output text file.
Row data is tab-separated. Sheet names are headers.
Metadata (filename, sheet count, row counts) is included.

Dropdown Logistics — Chaos -> Structured -> Automated
XLSX Converter v1.0 | 2026-03-07
"""

import os
import sys
import json
import argparse
import datetime
from openpyxl import load_workbook

# -----------------------------
# CONFIG
# -----------------------------
CANON_DIR = r"C:\Users\dexjr\99_DexUniverseArchive\00_Archive\DDL-Standards-Canon"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
LOG_FILE = os.path.join(SCRIPT_DIR, "dex-xlsx-log.jsonl")

# Max rows per sheet (0 = no limit)
MAX_ROWS = 0

# Max cell width before truncation (0 = no limit)
MAX_CELL_WIDTH = 500

# -----------------------------
# CONVERTER
# -----------------------------
def xlsx_to_text(filepath, preview=False):
    """Convert an xlsx file to structured text."""
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        print(f"  ERROR reading {filepath}: {e}")
        return None

    filename = os.path.basename(filepath)
    output_parts = []

    # Header
    output_parts.append(f"XLSX CONVERSION: {filename}")
    output_parts.append(f"Source: {filepath}")
    output_parts.append(f"Converted: {datetime.datetime.now().isoformat()}")
    output_parts.append(f"Sheets: {len(wb.sheetnames)}")
    output_parts.append(f"Sheet names: {', '.join(wb.sheetnames)}")
    output_parts.append("=" * 60)

    total_rows = 0
    total_cells = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        output_parts.append(f"\n{'─' * 60}")
        output_parts.append(f"SHEET: {sheet_name}")
        output_parts.append(f"{'─' * 60}\n")

        row_count = 0
        for row in ws.iter_rows(values_only=True):
            if preview and row_count >= 50:
                output_parts.append(f"\n... [PREVIEW: showing first 50 of {ws.max_row} rows]")
                break

            # Convert each cell to string
            cells = []
            for cell in row:
                if cell is None:
                    cells.append("")
                else:
                    val = str(cell)
                    if MAX_CELL_WIDTH > 0 and len(val) > MAX_CELL_WIDTH:
                        val = val[:MAX_CELL_WIDTH] + "..."
                    cells.append(val)

            # Skip completely empty rows
            if all(c == "" for c in cells):
                continue

            output_parts.append("\t".join(cells))
            row_count += 1
            total_cells += len(cells)

        total_rows += row_count
        output_parts.append(f"\n[{sheet_name}: {row_count} rows]")

    wb.close()

    # Footer
    output_parts.append(f"\n{'=' * 60}")
    output_parts.append(f"TOTAL: {total_rows} rows, {total_cells} cells across {len(wb.sheetnames)} sheets")
    output_parts.append(f"{'=' * 60}")

    text = "\n".join(output_parts)

    return {
        "text": text,
        "filename": filename,
        "sheets": len(wb.sheetnames),
        "sheet_names": wb.sheetnames,
        "total_rows": total_rows,
        "total_cells": total_cells,
        "char_count": len(text),
    }

# -----------------------------
# FILE OPERATIONS
# -----------------------------
def save_text(result, output_dir):
    """Save converted text to a file."""
    os.makedirs(output_dir, exist_ok=True)
    name = os.path.splitext(result["filename"])[0]
    output_path = os.path.join(output_dir, f"{name}_xlsx.txt")

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(result["text"])

    print(f"  Saved: {output_path} ({result['char_count']:,} chars)")
    return output_path

def ingest_text(result):
    """Save to canon folder for corpus ingestion."""
    os.makedirs(CANON_DIR, exist_ok=True)
    name = os.path.splitext(result["filename"])[0]
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(CANON_DIR, f"XLSX_{name}_{ts}.txt")

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(result["text"])

    print(f"  Ingested: {output_path} ({result['char_count']:,} chars)")
    return output_path

# -----------------------------
# BATCH PROCESSING
# -----------------------------
def find_xlsx_files(folder):
    """Find all .xlsx files in a folder (non-recursive)."""
    files = []
    for filename in os.listdir(folder):
        if filename.endswith(".xlsx") and not filename.startswith("~$"):
            files.append(os.path.join(folder, filename))
    return files

def find_xlsx_recursive(folder):
    """Find all .xlsx files recursively."""
    files = []
    for root, dirs, filenames in os.walk(folder):
        for filename in filenames:
            if filename.endswith(".xlsx") and not filename.startswith("~$"):
                files.append(os.path.join(root, filename))
    return files

# -----------------------------
# LOGGING
# -----------------------------
def log_conversion(filepath, result, action):
    entry = {
        "timestamp": datetime.datetime.now().isoformat(),
        "source": filepath,
        "filename": result["filename"],
        "sheets": result["sheets"],
        "total_rows": result["total_rows"],
        "total_cells": result["total_cells"],
        "char_count": result["char_count"],
        "action": action,
    }
    try:
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(json.dumps(entry, ensure_ascii=False) + "\n")
    except:
        pass

# -----------------------------
# MAIN
# -----------------------------
def main():
    parser = argparse.ArgumentParser(description="Dex Jr XLSX Converter v1.0")
    parser.add_argument("path", help="Path to .xlsx file or folder")
    parser.add_argument("--all", action="store_true", help="Convert all .xlsx in folder")
    parser.add_argument("--recursive", action="store_true", help="Search subfolders too")
    parser.add_argument("--ingest", action="store_true", help="Save to canon for ingestion")
    parser.add_argument("--save", default=None, help="Save to specified folder")
    parser.add_argument("--preview", action="store_true", help="Show first 50 rows per sheet")

    args = parser.parse_args()

    # Single file mode
    if not args.all and os.path.isfile(args.path):
        print(f"\n  Converting: {args.path}")
        result = xlsx_to_text(args.path, preview=args.preview)
        if not result:
            return

        print(f"  Sheets: {result['sheets']} ({', '.join(result['sheet_names'])})")
        print(f"  Rows: {result['total_rows']:,}")
        print(f"  Cells: {result['total_cells']:,}")
        print(f"  Text: {result['char_count']:,} chars")

        if args.ingest:
            ingest_text(result)
            log_conversion(args.path, result, "ingest")
        elif args.save:
            save_text(result, args.save)
            log_conversion(args.path, result, "save")
        else:
            # Display
            print(f"\n{'─' * 60}")
            preview_text = result["text"][:3000]
            print(preview_text)
            if len(result["text"]) > 3000:
                print(f"\n... [{result['char_count'] - 3000:,} more chars]")
            print(f"{'─' * 60}")
            log_conversion(args.path, result, "display")

        return

    # Batch mode
    folder = args.path
    if not os.path.isdir(folder):
        print(f"  ERROR: {folder} is not a directory")
        return

    if args.recursive:
        files = find_xlsx_recursive(folder)
    else:
        files = find_xlsx_files(folder)

    if not files:
        print(f"  No .xlsx files found in {folder}")
        return

    print(f"\n  Found {len(files)} .xlsx file(s)")
    print(f"  {'=' * 60}")

    total_rows = 0
    total_chars = 0
    converted = 0

    for i, filepath in enumerate(files):
        filename = os.path.basename(filepath)
        print(f"\n  [{i+1}/{len(files)}] {filename}...", end=" ")

        result = xlsx_to_text(filepath, preview=args.preview)
        if not result:
            continue

        print(f"{result['sheets']} sheets, {result['total_rows']:,} rows, {result['char_count']:,} chars")

        if args.ingest:
            ingest_text(result)
            log_conversion(filepath, result, "ingest")
        elif args.save:
            save_text(result, args.save)
            log_conversion(filepath, result, "save")
        else:
            log_conversion(filepath, result, "scan")

        total_rows += result["total_rows"]
        total_chars += result["char_count"]
        converted += 1

    print(f"\n  {'=' * 60}")
    print(f"  Converted: {converted} files")
    print(f"  Total rows: {total_rows:,}")
    print(f"  Total text: {total_chars:,} chars")
    if args.ingest:
        print(f"  All files saved to canon for ingestion")
    print(f"  {'=' * 60}\n")

if __name__ == "__main__":
    main()
