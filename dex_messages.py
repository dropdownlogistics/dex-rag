#!/usr/bin/env python3
"""
dex_messages.py — Automated iMessage export parser for StatCheck.

Parses iMazing CSV exports from 597+ conversation folders into a
unified merged file matching the DexMessageMerged_All schema.

Usage:
  python dex_messages.py                        # full parse, all folders
  python dex_messages.py --dry-run              # count files, preview
  python dex_messages.py --stats                # summary without file output
  python dex_messages.py --contact "Emily"      # single contact
  python dex_messages.py --since 2025-05-22     # only newer messages
  python dex_messages.py --output FILE          # custom output path

Dropdown Logistics — Chaos -> Structured -> Automated
"""

import argparse
import csv
import hashlib
import json
import os
import re
import sys
from datetime import datetime
from typing import Optional

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# ── Config ────────────────────────────────────────────────────────────────────

MESSAGES_DIR = r"C:\Users\dkitc\iCloudDrive\Documents\02_Dex\03_iMazingExports\iPhone\Messages"
V1_PATH = r"C:\Users\dkitc\iCloudDrive\Documents\02_Dex\00_Archive\DexKit_v4.1_Archive\17_Messages\DexMessageMerged_All_v1.0.xlsx"
DASHBOARD_PATH = r"C:\Users\dkitc\iCloudDrive\Documents\02_Dex\00_Archive\DexKit_v4.1_Archive\17_Messages\DexDashMaster_v1.0.xlsx"
DEFAULT_OUTPUT = r"C:\Users\dkitc\iCloudDrive\Documents\02_Dex\00_Archive\DexKit_v4.1_Archive\17_Messages\DexMessageMerged_All_v2.0.xlsx"
CLASSIFICATIONS_CACHE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "dex-rag-scratch", "msg_classifications.json")

ERA_BREAKPOINT = datetime(2025, 5, 1)

# Operator identifiers — messages FROM these are "Sent"
OPERATOR_IDS = {"dave", "dave kitchens", "d.k. hale", "me"}


# ── Helpers ───────────────────────────────────────────────────────────────────

def get_era(ts: datetime) -> str:
    return "PreDex" if ts < ERA_BREAKPOINT else "DexEra"


def clean_text(text: str) -> str:
    if not text:
        return ""
    t = text.strip()
    t = t.replace("\x00", "")
    t = re.sub(r"\n{3,}", "\n\n", t)
    return t


def word_count(text: str) -> int:
    return len(text.split()) if text else 0


def char_count(text: str) -> int:
    return len(text) if text else 0


def char_count_no_spaces(text: str) -> int:
    return len(text.replace(" ", "")) if text else 0


def msg_hash(filename: str, timestamp: str, text: str) -> str:
    key = f"{filename}|{timestamp}|{(text or '')[:50]}"
    return hashlib.md5(key.encode("utf-8", errors="replace")).hexdigest()


def extract_contact_name(folder_name: str) -> str:
    """Extract contact name from folder like '2023-04-03 10 52 56 - Tidal Wave'"""
    parts = folder_name.split(" - ", 1)
    if len(parts) > 1:
        return parts[1].strip()
    return folder_name


def derive_filename(contact_name: str) -> str:
    """Convert contact name to filename matching v1.0 format (e.g., 'Todd_Kitchens.txt')"""
    safe = re.sub(r"[^\w\s]", "", contact_name)
    safe = re.sub(r"\s+", "_", safe.strip())
    return f"{safe}.txt" if safe else "Unknown.txt"


def determine_direction(sender_name: str, msg_type: str) -> str:
    """Determine Sent/Received from sender name and iMazing Type field."""
    if msg_type and msg_type.strip().lower() == "outgoing":
        return "Sent"
    if msg_type and msg_type.strip().lower() == "incoming":
        return "Received"
    if sender_name:
        if sender_name.strip().lower() in OPERATOR_IDS:
            return "Sent"
    return "Received"


def load_classifications() -> dict:
    """Load pre-existing contact classifications from dashboard."""
    if os.path.exists(CLASSIFICATIONS_CACHE):
        try:
            with open(CLASSIFICATIONS_CACHE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


# ── CSV Parser ────────────────────────────────────────────────────────────────

def parse_folder(folder_path: str, folder_name: str,
                 since: Optional[datetime] = None,
                 contact_filter: Optional[str] = None) -> list[dict]:
    """Parse a single iMazing export folder. Returns list of message dicts."""
    contact_name = extract_contact_name(folder_name)

    if contact_filter and contact_filter.lower() not in contact_name.lower():
        return []

    csvs = [f for f in os.listdir(folder_path) if f.endswith(".csv")]
    if not csvs:
        return []

    csv_path = os.path.join(folder_path, csvs[0])
    filename = derive_filename(contact_name)
    messages = []

    try:
        with open(csv_path, encoding="utf-8", errors="replace", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                # Parse timestamp
                date_str = row.get("Message Date", "").strip()
                if not date_str:
                    continue
                try:
                    ts = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    try:
                        ts = datetime.strptime(date_str[:19], "%Y-%m-%dT%H:%M:%S")
                    except ValueError:
                        continue

                if since and ts <= since:
                    continue

                # Extract fields
                text = row.get("Text", "")
                attachment = row.get("Attachment", "")
                sender = row.get("Sender Name", "") or row.get("Sender ID", "")
                msg_type = row.get("Type", "")

                cleaned = clean_text(text)
                if not cleaned and attachment:
                    cleaned = "(Attachment)"

                direction = determine_direction(sender, msg_type)

                messages.append({
                    "filename": filename,
                    "contact_name": contact_name,
                    "timestamp": ts,
                    "date": ts.date(),
                    "speaker": sender.strip() if sender else contact_name,
                    "message_cleaned": cleaned,
                    "direction": direction,
                    "word_count": word_count(cleaned),
                    "char_count": char_count(cleaned),
                    "char_count_no_spaces": char_count_no_spaces(cleaned),
                    "era": get_era(ts),
                })

    except Exception as e:
        print(f"  [ERROR] {folder_name}: {e}", file=sys.stderr)

    return messages


# ── Main pipeline ─────────────────────────────────────────────────────────────

def run_parse(dry_run=False, stats_only=False, output_path=None,
              since=None, contact_filter=None):

    if not os.path.isdir(MESSAGES_DIR):
        print(f"  ERROR: Messages directory not found: {MESSAGES_DIR}")
        return

    folders = sorted([f for f in os.listdir(MESSAGES_DIR)
                      if os.path.isdir(os.path.join(MESSAGES_DIR, f))])

    print(f"\n{'='*60}")
    print(f"  DEX MESSAGES — iMessage Export Parser")
    print(f"{'='*60}")
    print(f"  Source:  {MESSAGES_DIR}")
    print(f"  Folders: {len(folders)}")
    if since:
        print(f"  Since:   {since}")
    if contact_filter:
        print(f"  Filter:  {contact_filter}")
    if dry_run:
        print(f"  Mode:    DRY RUN")
    elif stats_only:
        print(f"  Mode:    STATS ONLY (no file output)")
    print(f"{'='*60}\n")

    if dry_run:
        # Just count CSVs and estimate
        csv_count = 0
        empty = 0
        for folder in folders:
            fpath = os.path.join(MESSAGES_DIR, folder)
            csvs = [f for f in os.listdir(fpath) if f.endswith(".csv")]
            if csvs:
                csv_count += 1
            else:
                empty += 1
        print(f"  Folders with CSV: {csv_count}")
        print(f"  Empty folders:    {empty}")
        print(f"  Estimated messages: ~{csv_count * 400} (avg 400/contact)")
        return

    # Parse all folders
    all_messages = []
    seen_hashes = set()
    parsed = 0
    skipped = 0
    dupes = 0
    errors = []

    for i, folder in enumerate(folders):
        fpath = os.path.join(MESSAGES_DIR, folder)
        msgs = parse_folder(fpath, folder, since=since, contact_filter=contact_filter)

        for msg in msgs:
            h = msg_hash(msg["filename"], str(msg["timestamp"]), msg["message_cleaned"])
            if h in seen_hashes:
                dupes += 1
                continue
            seen_hashes.add(h)
            all_messages.append(msg)

        if msgs:
            parsed += 1
        elif not contact_filter:
            skipped += 1

        if (i + 1) % 100 == 0:
            print(f"  [{i+1}/{len(folders)}] {len(all_messages)} messages parsed...")

    # Sort by timestamp
    all_messages.sort(key=lambda m: m["timestamp"])

    # Assign sequential MessageIDs
    for i, msg in enumerate(all_messages, 1):
        msg["message_id"] = i

    # Load classifications for stats
    classifications = load_classifications()

    # Compute stats
    era_counts = {}
    rel_counts = {}
    contact_counts = {}
    dir_counts = {"Sent": 0, "Received": 0}

    for msg in all_messages:
        era_counts[msg["era"]] = era_counts.get(msg["era"], 0) + 1
        dir_counts[msg["direction"]] = dir_counts.get(msg["direction"], 0) + 1

        cn = msg["contact_name"]
        contact_counts[cn] = contact_counts.get(cn, 0) + 1

        fn = msg["filename"]
        cls = classifications.get(fn, {})
        rel = cls.get("relationship", "Unclassified")
        if rel:
            rel_counts[rel] = rel_counts.get(rel, 0) + 1

    # Date range
    min_date = all_messages[0]["timestamp"] if all_messages else None
    max_date = all_messages[-1]["timestamp"] if all_messages else None

    # Top 10
    top10 = sorted(contact_counts.items(), key=lambda x: -x[1])[:10]

    # Print stats
    print(f"\n{'='*60}")
    print(f"  DEX MESSAGES — Parse Complete")
    print(f"{'='*60}")
    print(f"  Folders scanned:  {len(folders)}")
    print(f"  Folders parsed:   {parsed} ({skipped} empty/filtered)")
    print(f"  Total messages:   {len(all_messages):,}")
    print(f"  Duplicates:       {dupes:,}")
    if min_date and max_date:
        print(f"  Date range:       {min_date.strftime('%Y-%m-%d')} to {max_date.strftime('%Y-%m-%d')}")

    print(f"\n  By Direction:")
    for d, c in sorted(dir_counts.items()):
        pct = c / len(all_messages) * 100 if all_messages else 0
        print(f"    {d:10s} {c:>8,}  ({pct:.1f}%)")

    print(f"\n  By Era:")
    for era in ["PreDex", "DexEra"]:
        c = era_counts.get(era, 0)
        print(f"    {era:10s} {c:>8,}")

    if rel_counts:
        print(f"\n  By Relationship:")
        for rel, c in sorted(rel_counts.items(), key=lambda x: -x[1]):
            print(f"    {rel:15s} {c:>8,}")

    print(f"\n  Top 10 by message count:")
    for name, count in top10:
        print(f"    {name:30s} {count:>6,}")

    # Reconciliation against v1.0
    if os.path.exists(V1_PATH):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(V1_PATH, read_only=True, data_only=True)
            ws = wb["MessageLogFull"]
            v1_count = ws.max_row - 2  # subtract header rows
            wb.close()
            delta = len(all_messages) - v1_count
            print(f"\n  RECONCILIATION vs v1.0:")
            print(f"    v1.0 messages:   {v1_count:>8,}")
            print(f"    v2.0 messages:   {len(all_messages):>8,}")
            print(f"    Delta:           {delta:>+8,}")
        except Exception as e:
            print(f"\n  RECONCILIATION: Could not read v1.0 ({e})")

    # Write output
    if not stats_only and all_messages:
        out_path = output_path or DEFAULT_OUTPUT
        print(f"\n  Writing to: {out_path}")

        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "MessageLogFull"

            # Headers (row 2, matching v1.0 layout — row 1 blank)
            headers = ["", "MessageID", "Filename", "Timestamp", "Date",
                        "Speaker", "Message_Cleaned", "Direction",
                        "WordCount", "CharCount", "CharCountNoSpaces", "Era"]

            # Row 1 blank, row 2 headers
            ws.append([])
            ws.append(headers)

            # Data rows
            for msg in all_messages:
                ws.append([
                    None,
                    msg["message_id"],
                    msg["filename"],
                    msg["timestamp"],
                    msg["date"],
                    msg["speaker"],
                    msg["message_cleaned"],
                    msg["direction"],
                    msg["word_count"],
                    msg["char_count"],
                    msg["char_count_no_spaces"],
                    msg["era"],
                ])

            wb.save(out_path)
            size_mb = os.path.getsize(out_path) / 1024 / 1024
            print(f"  Written: {size_mb:.1f} MB")
        except Exception as e:
            print(f"  ERROR writing output: {e}")

    # Unclassified contacts
    new_contacts = []
    for name in contact_counts:
        fn = derive_filename(name)
        if fn not in classifications:
            new_contacts.append((name, contact_counts[name]))
    if new_contacts:
        new_contacts.sort(key=lambda x: -x[1])
        print(f"\n  Unclassified contacts ({len(new_contacts)}):")
        for name, count in new_contacts[:15]:
            print(f"    {name:30s} {count:>6,} msgs")
        if len(new_contacts) > 15:
            print(f"    ... and {len(new_contacts) - 15} more")

    print(f"\n{'='*60}\n")


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(description="Dex Messages — iMessage Export Parser")
    p.add_argument("--dry-run", action="store_true", help="Count files, preview only")
    p.add_argument("--stats", action="store_true", help="Summary only, no file output")
    p.add_argument("--output", default=None, help="Custom output path")
    p.add_argument("--since", default=None, help="Only messages after this date (YYYY-MM-DD)")
    p.add_argument("--contact", default=None, help="Filter to single contact name")
    args = p.parse_args()

    since = None
    if args.since:
        try:
            since = datetime.strptime(args.since, "%Y-%m-%d")
        except ValueError:
            print(f"  ERROR: Invalid date format: {args.since} (use YYYY-MM-DD)")
            return

    run_parse(
        dry_run=args.dry_run,
        stats_only=args.stats or args.dry_run,
        output_path=args.output,
        since=since,
        contact_filter=args.contact,
    )


if __name__ == "__main__":
    main()
